"""
Advanced Cleavage Mapping Automation
Complete implementation with bidirectional panel linking
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict
from typing import Dict, List, Tuple, Optional
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pandas as pd

class AdvancedCleavageMapper:
    """
    Complete automation for peptide cleavage mapping analysis
    Handles both N-terminal (left) and C-terminal (right) grouping with formula linkage
    """
    
    def __init__(self, workbook_path: str):
        self.wb = openpyxl.load_workbook(workbook_path)
        self.intensity_start_col = 3  # Column C
        self.intensity_end_col = 9     # Column I
        self.num_samples = 7
        
    def analyze_sequence_structure(self, sequences: List[Dict], reference: str) -> Dict:
        """
        Analyze sequences to determine truncation patterns
        Returns mapping of sequences to their truncation positions
        """
        analysis = {
            'n_terminal_groups': defaultdict(list),
            'c_terminal_groups': defaultdict(list),
            'sequence_mapping': {}
        }
        
        for seq_data in sequences:
            clean = seq_data['clean']
            
            # Determine N-terminal truncation position
            if clean in reference:
                n_term_pos = reference.index(clean)
            else:
                n_term_pos = None
            
            # Determine C-terminal truncation position
            if clean in reference:
                c_term_pos = reference.index(clean) + len(clean)
            else:
                c_term_pos = None
            
            seq_data['n_term_position'] = n_term_pos
            seq_data['c_term_position'] = c_term_pos
            
            # Group by cleavage residue
            if seq_data['left_cleavage']:
                analysis['n_terminal_groups'][seq_data['left_cleavage']].append(seq_data)
            if seq_data['right_cleavage']:
                analysis['c_terminal_groups'][seq_data['right_cleavage']].append(seq_data)
        
        return analysis
    
    def build_left_panel_structure(self, n_terminal_groups: Dict) -> List[Dict]:
        """
        Build left panel structure with row assignments
        Each group gets contiguous rows, followed by spacing
        """
        structure = []
        current_row = 3  # Start after headers and template
        
        for residue in sorted(n_terminal_groups.keys()):
            group_seqs = n_terminal_groups[residue]
            
            # Sort sequences by length for consistent display
            sorted_seqs = sorted(group_seqs, key=lambda x: len(x['clean']))
            
            group_info = {
                'residue': residue,
                'start_row': current_row,
                'sequences': []
            }
            
            for seq in sorted_seqs:
                seq_row = {
                    'row': current_row,
                    'data': seq
                }
                group_info['sequences'].append(seq_row)
                current_row += 1
            
            group_info['end_row'] = current_row - 1
            structure.append(group_info)
            
            # Add spacing between groups
            current_row += 1
        
        return structure
    
    def build_right_panel_structure(self, c_terminal_groups: Dict, sequences: List[Dict]) -> List[Dict]:
        """
        Build right panel structure showing C-terminal truncations
        Organized by progressive shortening from C-terminus
        """
        structure = []
        current_row = 3
        
        # Group all sequences by their C-terminal position
        by_c_term_length = defaultdict(list)
        for seq in sequences:
            by_c_term_length[len(seq['clean'])].append(seq)
        
        # Sort by length (longest first = least C-terminal truncation)
        for length in sorted(by_c_term_length.keys(), reverse=True):
            group_seqs = by_c_term_length[length]
            
            group_info = {
                'length': length,
                'start_row': current_row,
                'sequences': []
            }
            
            for seq in group_seqs:
                seq_row = {
                    'row': current_row,
                    'data': seq
                }
                group_info['sequences'].append(seq_row)
                current_row += 1
            
            group_info['end_row'] = current_row - 1
            structure.append(group_info)
        
        return structure
    
    def generate_panel_linkage(self, left_structure: List[Dict], right_structure: List[Dict]) -> Dict[int, Tuple[int, int]]:
        """
        Determine which right panel rows correspond to each left panel group
        Returns mapping: left_row -> (right_start_row, right_end_row)
        """
        linkage = {}
        
        for left_group in left_structure:
            # Extract sequences in this left group
            left_seqs = [s['data']['clean'] for s in left_group['sequences']]
            
            # Find matching sequences in right panel
            matching_right_rows = []
            for right_group in right_structure:
                for seq_row in right_group['sequences']:
                    if seq_row['data']['clean'] in left_seqs:
                        matching_right_rows.append(seq_row['row'])
            
            if matching_right_rows:
                linkage[left_group['start_row']] = (min(matching_right_rows), max(matching_right_rows))
        
        return linkage
    
    def write_processed_worksheet(self, 
                                  raw_data: Dict, 
                                  output_sheet_name: str,
                                  sample_labels: Optional[List[str]] = None):
        """
        Generate complete processed worksheet with all formulas
        """
        sequences = raw_data['sequences']
        reference = raw_data['reference']
        
        # Analyze structure
        analysis = self.analyze_sequence_structure(sequences, reference)
        
        # Build panel structures
        left_structure = self.build_left_panel_structure(analysis['n_terminal_groups'])
        right_structure = self.build_right_panel_structure(analysis['c_terminal_groups'], sequences)
        
        # Determine linkage
        linkage = self.generate_panel_linkage(left_structure, right_structure)
        
        # Create/clear worksheet
        if output_sheet_name in self.wb.sheetnames:
            ws = self.wb[output_sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = self.wb.create_sheet(output_sheet_name)
        
        # Write headers
        self._write_complete_headers(ws, sample_labels)
        
        # Write left panel data
        for group in left_structure:
            for seq_row in group['sequences']:
                row = seq_row['row']
                data = seq_row['data']
                
                ws.cell(row, 1).value = data['number']  # Number
                ws.cell(row, 2).value = data['clean']   # Sequence
                
                # Intensity values
                for i, intensity in enumerate(data['intensities']):
                    ws.cell(row, 3 + i).value = intensity
                
                ws.cell(row, 12).value = group['residue']  # Left cleavage indicator
        
        # Write right panel data
        for group in right_structure:
            for seq_row in group['sequences']:
                row = seq_row['row']
                data = seq_row['data']
                
                ws.cell(row, 19).value = data['clean']  # Sequence in right panel
                
                # Intensity values
                for i, intensity in enumerate(data['intensities']):
                    ws.cell(row, 20 + i).value = intensity
        
        # Add formulas
        self._write_formulas(ws, left_structure, right_structure, linkage)
        
        return ws
    
    def _write_complete_headers(self, ws, sample_labels: Optional[List[str]] = None):
        """Write comprehensive headers for both panels"""
        if sample_labels is None:
            sample_labels = [f'Sample_{i}' for i in range(1, 8)]
        
        # Left panel
        ws.cell(1, 1).value = '#'
        ws.cell(1, 2).value = 'Sequence'
        for i, label in enumerate(sample_labels):
            ws.cell(1, 3 + i).value = label
        
        # Calculation columns
        ws.cell(1, 12).value = 'left'
        ws.cell(1, 13).value = 'right'
        ws.cell(1, 14).value = 'sum'
        ws.cell(1, 15).value = 'percentage'
        
        # Right panel
        ws.cell(1, 18).value = '#'
        ws.cell(1, 19).value = 'Sequence'
        for i, label in enumerate(sample_labels):
            ws.cell(1, 20 + i).value = label
    
    def _write_formulas(self, ws, left_structure: List[Dict], right_structure: List[Dict], linkage: Dict):
        """
        Write all calculation formulas
        """
        # Determine total row (last group's end + 1)
        total_row = max(g['end_row'] for g in left_structure) + 1
        
        # Write formulas for each left panel group
        for group in left_structure:
            summary_row = group['end_row'] + 1  # Row after group for summary
            start = group['start_row']
            end = group['end_row']
            
            # Column M: Sum of left panel intensities
            ws.cell(summary_row, 13).value = f"=SUM(C{start}:I{end})"
            
            # Column N: Sum of corresponding right panel rows
            if start in linkage:
                right_start, right_end = linkage[start]
                ws.cell(summary_row, 14).value = f"=SUM(T{right_start}:Z{right_end})"
            else:
                ws.cell(summary_row, 14).value = 0
            
            # Column O: Total
            ws.cell(summary_row, 15).value = f"=M{summary_row}+N{summary_row}"
            
            # Column P: Percentage
            ws.cell(summary_row, 16).value = f"=(O{summary_row}/$O${total_row})*100"
        
        # Write total row formulas
        first_summary_row = left_structure[0]['end_row'] + 1
        last_summary_row = left_structure[-1]['end_row'] + 1
        
        ws.cell(total_row, 13).value = f"=SUM(M{first_summary_row}:M{last_summary_row})"
        ws.cell(total_row, 14).value = f"=SUM(N{first_summary_row}:N{last_summary_row})"
        ws.cell(total_row, 15).value = f"=SUM(O{first_summary_row}:O{last_summary_row})"
        ws.cell(total_row, 16).value = "100"  # Total percentage
    
    def parse_raw_worksheet(self, sheet_name: str) -> Dict:
        """Parse raw data from worksheet"""
        ws = self.wb[sheet_name]
        
        # Extract reference sequence (row 4)
        reference = ws.cell(4, 2).value
        if not reference:
            raise ValueError(f"No reference sequence found in {sheet_name} row 4")
        
        sequences = []
        pattern = re.compile(r'\(([A-Z])\)')
        
        for row in range(5, ws.max_row + 1):
            seq = ws.cell(row, 2).value
            if not seq:
                continue
            
            # Parse cleavage notation
            cleavages = pattern.findall(seq)
            clean_seq = re.sub(r'[()]', '', seq)
            
            # Extract intensities
            intensities = []
            has_data = False
            for col in range(self.intensity_start_col, self.intensity_end_col + 1):
                val = ws.cell(row, col).value
                if val and val > 0:
                    has_data = True
                intensities.append(val if val else 0)
            
            if has_data:
                sequences.append({
                    'number': ws.cell(row, 1).value,
                    'original': seq,
                    'clean': clean_seq,
                    'left_cleavage': cleavages[0] if len(cleavages) > 0 else None,
                    'right_cleavage': cleavages[1] if len(cleavages) > 1 else None,
                    'intensities': intensities
                })
        
        return {
            'reference': reference,
            'sequences': sequences
        }
    
    def process(self, input_sheet: str, output_sheet: str, sample_labels: Optional[List[str]] = None):
        """
        Complete processing pipeline
        """
        print(f"\nProcessing: {input_sheet} -> {output_sheet}")
        
        # Parse input
        raw_data = self.parse_raw_worksheet(input_sheet)
        print(f"  Reference: {raw_data['reference'][:50]}...")
        print(f"  Sequences: {len(raw_data['sequences'])}")
        
        # Generate output
        ws = self.write_processed_worksheet(raw_data, output_sheet, sample_labels)
        print(f"  Generated: {ws.max_row} rows")
        
        return ws
    
    def save(self, output_path: Optional[str] = None):
        """Save workbook"""
        if output_path:
            self.wb.save(output_path)
            print(f"\nSaved: {output_path}")
        else:
            # Default to saving with processed suffix
            default_path = "processed_cleavage_mapper_output.xlsx"
            self.wb.save(default_path)
            print(f"\nSaved: {default_path}")
    
    def create_positional_intensity_heatmap(self, 
                                           raw_data: Dict, 
                                           sample_labels: Optional[List[str]] = None,
                                           output_path: str = "positional_intensity_heatmap.png",
                                           figsize: Tuple[int, int] = (14, 10)):
        """
        Create a heatmap showing intensities by amino acid position in the reference sequence
        Y-axis shows amino acid positions, X-axis shows samples
        """
        sequences = raw_data['sequences']
        reference = raw_data['reference']
        
        if sample_labels is None:
            sample_labels = [f'Sample_{i}' for i in range(1, self.num_samples + 1)]
        
        # Create position-based intensity matrix
        # Each position in reference sequence gets aggregated intensities
        position_intensities = np.zeros((len(reference), len(sample_labels)))
        position_counts = np.zeros((len(reference), len(sample_labels)))
        
        for seq_data in sequences:
            clean_seq = seq_data['clean']
            intensities = seq_data['intensities']
            
            # Find where this sequence maps in the reference
            if clean_seq in reference:
                start_pos = reference.index(clean_seq)
                end_pos = start_pos + len(clean_seq)
                
                # Add intensities to each position covered by this sequence
                for pos in range(start_pos, end_pos):
                    for sample_idx, intensity in enumerate(intensities):
                        if intensity and intensity > 0:
                            position_intensities[pos, sample_idx] += intensity
                            position_counts[pos, sample_idx] += 1
        
        # Create position labels (amino acid + position number)
        position_labels = [f'{reference[i]}{i+1}' for i in range(len(reference))]
        
        # Create DataFrame for easier handling
        df = pd.DataFrame(position_intensities, 
                         index=position_labels, 
                         columns=sample_labels)
        
        # Filter out positions with no data
        df = df.loc[(df.sum(axis=1) > 0)]
        
        if df.empty:
            print("⚠ No positional data to plot")
            return None
        
        # Create the heatmap
        plt.figure(figsize=figsize)
        
        # Use log scale for better visualization
        log_data = np.log10(df + 1)
        
        sns.heatmap(log_data, 
                   annot=False,
                   cmap='plasma',
                   cbar_kws={'label': 'Log10(Intensity + 1)'},
                   xticklabels=True,
                   yticklabels=True)
        
        plt.title(f'Peptide Position Intensity Heatmap\n({len(df)} positions with data)')
        plt.xlabel('Samples')
        plt.ylabel('Amino Acid Position')
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0, fontsize=8)
        plt.tight_layout()
        
        # Save the plot
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"✓ Positional heatmap saved to: {output_path}")
        print(f"✓ Positions with data: {len(df)}/{len(reference)}")
        
        return plt.gcf()

    def create_intensity_heatmap(self, 
                                raw_data: Dict, 
                                sample_labels: Optional[List[str]] = None,
                                output_path: str = "intensity_heatmap.png",
                                figsize: Tuple[int, int] = (12, 8),
                                top_n: Optional[int] = None):
        """
        Create a heatmap showing intensities for each peptide across samples
        
        Args:
            raw_data: Parsed raw data from parse_raw_worksheet
            sample_labels: Labels for the samples
            output_path: Path to save the heatmap image
            figsize: Figure size (width, height)
            top_n: Show only top N peptides by total intensity (None for all)
        """
        sequences = raw_data['sequences']
        
        if sample_labels is None:
            sample_labels = [f'Sample_{i}' for i in range(1, self.num_samples + 1)]
        
        # Create data matrix
        peptide_names = []
        intensity_matrix = []
        
        for seq in sequences:
            peptide_names.append(seq['clean'][:20] + ('...' if len(seq['clean']) > 20 else ''))
            intensity_matrix.append(seq['intensities'])
        
        # Convert to numpy array and pandas DataFrame
        intensity_array = np.array(intensity_matrix)
        df = pd.DataFrame(intensity_array, 
                         index=peptide_names, 
                         columns=sample_labels)
        
        # Filter to top N if specified
        if top_n and top_n < len(df):
            total_intensities = df.sum(axis=1)
            top_peptides = total_intensities.nlargest(top_n).index
            df = df.loc[top_peptides]
        
        # Create the heatmap
        plt.figure(figsize=figsize)
        
        # Use log scale for better visualization of wide intensity ranges
        # Add 1 to avoid log(0) issues
        log_data = np.log10(df + 1)
        
        sns.heatmap(log_data, 
                   annot=False,  # Don't annotate due to space constraints
                   cmap='viridis',
                   cbar_kws={'label': 'Log10(Intensity + 1)'},
                   xticklabels=True,
                   yticklabels=True)
        
        plt.title(f'Peptide Intensity Heatmap\n({len(df)} peptides across {len(sample_labels)} samples)')
        plt.xlabel('Samples')
        plt.ylabel('Peptides')
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0, fontsize=8)
        plt.tight_layout()
        
        # Save the plot
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"✓ Heatmap saved to: {output_path}")
        
        # Show basic statistics
        print(f"✓ Heatmap statistics:")
        print(f"  - Peptides: {len(df)}")
        print(f"  - Samples: {len(sample_labels)}")
        print(f"  - Max intensity: {df.max().max():,.0f}")
        print(f"  - Min intensity: {df.min().min():,.0f}")
        print(f"  - Mean intensity: {df.mean().mean():,.0f}")
        
        return plt.gcf()
    
    def create_cleavage_summary_plot(self, 
                                   raw_data: Dict,
                                   sample_labels: Optional[List[str]] = None,
                                   output_path: str = "cleavage_summary.png",
                                   figsize: Tuple[int, int] = (10, 6)):
        """
        Create a summary plot showing cleavage patterns
        """
        sequences = raw_data['sequences']
        
        if sample_labels is None:
            sample_labels = [f'Sample_{i}' for i in range(1, self.num_samples + 1)]
        
        # Analyze cleavage patterns
        analysis = self.analyze_sequence_structure(sequences, raw_data['reference'])
        
        # Prepare data for plotting
        n_term_data = {}
        c_term_data = {}
        
        for residue, seqs in analysis['n_terminal_groups'].items():
            total_intensity = sum(sum(seq['intensities']) for seq in seqs)
            n_term_data[residue] = total_intensity
        
        for residue, seqs in analysis['c_terminal_groups'].items():
            total_intensity = sum(sum(seq['intensities']) for seq in seqs)
            c_term_data[residue] = total_intensity
        
        # Create subplots
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=figsize)
        
        # N-terminal cleavage plot
        if n_term_data:
            residues = list(n_term_data.keys())
            intensities = list(n_term_data.values())
            ax1.bar(residues, intensities, color='skyblue', alpha=0.7)
            ax1.set_title('N-terminal Cleavage Patterns')
            ax1.set_xlabel('Cleavage Residue')
            ax1.set_ylabel('Total Intensity')
            ax1.tick_params(axis='x', rotation=45)
        
        # C-terminal cleavage plot
        if c_term_data:
            residues = list(c_term_data.keys())
            intensities = list(c_term_data.values())
            ax2.bar(residues, intensities, color='lightcoral', alpha=0.7)
            ax2.set_title('C-terminal Cleavage Patterns')
            ax2.set_xlabel('Cleavage Residue')
            ax2.set_ylabel('Total Intensity')
            ax2.tick_params(axis='x', rotation=45)
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"✓ Cleavage summary plot saved to: {output_path}")
        
        return fig
    
    def create_visualizations(self, 
                            input_sheet: str, 
                            sample_labels: Optional[List[str]] = None,
                            output_prefix: str = "cleavage_analysis",
                            top_n_peptides: Optional[int] = 50):
        """
        Create all visualizations for a given worksheet
        
        Args:
            input_sheet: Name of the input worksheet
            sample_labels: Labels for the samples  
            output_prefix: Prefix for output filenames
            top_n_peptides: Number of top peptides to show in heatmap
        """
        print(f"\nCreating visualizations for: {input_sheet}")
        
        # Parse the data
        raw_data = self.parse_raw_worksheet(input_sheet)
        
        # Create traditional sequence heatmap
        heatmap_path = f"{output_prefix}_heatmap.png"
        self.create_intensity_heatmap(raw_data, sample_labels, heatmap_path, 
                                    top_n=top_n_peptides)
        
        # Create positional heatmap
        positional_path = f"{output_prefix}_positional_heatmap.png"
        self.create_positional_intensity_heatmap(raw_data, sample_labels, positional_path)
        
        # Create cleavage summary
        summary_path = f"{output_prefix}_cleavage_summary.png"
        self.create_cleavage_summary_plot(raw_data, sample_labels, summary_path)
        
        # Close plots to free memory
        plt.close('all')
        
        print(f"✓ All visualizations created with prefix: {output_prefix}")
        
        return [heatmap_path, positional_path, summary_path]
    
    def create_comprehensive_report(self, 
                                  conditions: List[Tuple[str, str]] = None,
                                  sample_labels: Optional[List[str]] = None,
                                  output_path: str = "comprehensive_cleavage_report.png",
                                  figsize: Tuple[int, int] = (20, 16)):
        """
        Create a comprehensive report showing all conditions in a single visualization
        
        Args:
            conditions: List of (worksheet_name, display_name) tuples
            sample_labels: Labels for the samples
            output_path: Path to save the report
            figsize: Figure size
        """
        if conditions is None:
            conditions = [
                ('100 mgd glucose', '100 mgd'),
                ('200 mgd glucose', '200 mgd'), 
                ('500 mgd glucose', '500 mgd')
            ]
        
        if sample_labels is None:
            sample_labels = [f'Sample_{i}' for i in range(1, self.num_samples + 1)]
        
        print(f"\n=== Creating Comprehensive Report ===")
        
        # Load data for all conditions
        all_data = {}
        for worksheet, display_name in conditions:
            if worksheet in self.wb.sheetnames:
                try:
                    raw_data = self.parse_raw_worksheet(worksheet)
                    all_data[display_name] = raw_data
                    print(f"✓ Loaded {display_name}: {len(raw_data['sequences'])} sequences")
                except Exception as e:
                    print(f"⚠ Could not load {worksheet}: {e}")
        
        if not all_data:
            print("✗ No data available for report")
            return None
        
        # Create comprehensive figure with multiple subplots
        fig = plt.figure(figsize=figsize)
        
        # Create dynamic grid layout based on number of conditions
        n_conditions = len(all_data)
        n_cols = min(n_conditions, 4)  # Max 4 columns
        
        gs = fig.add_gridspec(3, n_cols, height_ratios=[2, 2, 1], 
                             width_ratios=[1] * n_cols)
        
        # Top row: Positional heatmaps for each condition
        for idx, (condition_name, raw_data) in enumerate(all_data.items()):
            if idx >= n_cols:  # Skip if too many conditions
                continue
            ax = fig.add_subplot(gs[0, idx])
            
            # Create positional intensity matrix
            reference = raw_data['reference']
            sequences = raw_data['sequences']
            
            position_intensities = np.zeros((len(reference), len(sample_labels)))
            
            for seq_data in sequences:
                clean_seq = seq_data['clean']
                intensities = seq_data['intensities']
                
                if clean_seq in reference:
                    start_pos = reference.index(clean_seq)
                    end_pos = start_pos + len(clean_seq)
                    
                    for pos in range(start_pos, end_pos):
                        for sample_idx, intensity in enumerate(intensities):
                            if intensity and intensity > 0:
                                position_intensities[pos, sample_idx] += intensity
            
            # Filter positions with data and create labels
            has_data = position_intensities.sum(axis=1) > 0
            filtered_positions = position_intensities[has_data]
            position_labels = [f'{reference[i]}{i+1}' for i in range(len(reference)) if has_data[i]]
            
            if len(filtered_positions) > 0:
                log_data = np.log10(filtered_positions + 1)
                
                im = ax.imshow(log_data, cmap='plasma', aspect='auto')
                ax.set_title(f'{condition_name}\nPositional Intensities')
                ax.set_xlabel('Samples')
                ax.set_ylabel('AA Position')
                
                # Set ticks
                ax.set_xticks(range(len(sample_labels)))
                ax.set_xticklabels([s.replace('AspN_', '') for s in sample_labels], 
                                  rotation=45, fontsize=8)
                
                # Show every 5th position for readability
                step = max(1, len(position_labels) // 10)
                tick_positions = range(0, len(position_labels), step)
                ax.set_yticks(tick_positions)
                ax.set_yticklabels([position_labels[i] for i in tick_positions], fontsize=6)
            else:
                ax.text(0.5, 0.5, 'No Data', ha='center', va='center', transform=ax.transAxes)
                ax.set_title(f'{condition_name}\n(No Data)')
        
        # Middle row: Cleavage pattern comparisons
        middle_cols = min(3, n_cols)
        ax_n_term = fig.add_subplot(gs[1, 0])
        ax_c_term = fig.add_subplot(gs[1, 1]) if middle_cols > 1 else None
        ax_total = fig.add_subplot(gs[1, 2]) if middle_cols > 2 else fig.add_subplot(gs[1, 1])
        
        # Collect cleavage data
        n_term_comparison = {}
        c_term_comparison = {}
        total_intensities = {}
        
        for condition_name, raw_data in all_data.items():
            analysis = self.analyze_sequence_structure(raw_data['sequences'], raw_data['reference'])
            
            # N-terminal data
            n_term_data = {}
            for residue, seqs in analysis['n_terminal_groups'].items():
                total = sum(sum(seq['intensities']) for seq in seqs)
                n_term_data[residue] = total
            n_term_comparison[condition_name] = n_term_data
            
            # C-terminal data  
            c_term_data = {}
            for residue, seqs in analysis['c_terminal_groups'].items():
                total = sum(sum(seq['intensities']) for seq in seqs)
                c_term_data[residue] = total
            c_term_comparison[condition_name] = c_term_data
            
            # Total intensity
            total_intensities[condition_name] = sum(sum(seq['intensities']) for seq in raw_data['sequences'])
        
        # Plot N-terminal comparison
        all_n_residues = set()
        for data in n_term_comparison.values():
            all_n_residues.update(data.keys())
        
        if all_n_residues:
            x_pos = np.arange(len(all_n_residues))
            width = 0.25
            
            for idx, (condition, data) in enumerate(n_term_comparison.items()):
                values = [data.get(residue, 0) for residue in sorted(all_n_residues)]
                ax_n_term.bar(x_pos + idx * width, values, width, 
                             label=condition, alpha=0.8)
            
            ax_n_term.set_xlabel('N-terminal Cleavage Residue')
            ax_n_term.set_ylabel('Total Intensity')
            ax_n_term.set_title('N-terminal Cleavage Comparison')
            ax_n_term.set_xticks(x_pos + width)
            ax_n_term.set_xticklabels(sorted(all_n_residues))
            ax_n_term.legend(fontsize=8)
        
        # Plot C-terminal comparison (only if we have space)
        if ax_c_term is not None:
            all_c_residues = set()
            for data in c_term_comparison.values():
                all_c_residues.update(data.keys())
            
            if all_c_residues:
                x_pos = np.arange(len(all_c_residues))
                
                for idx, (condition, data) in enumerate(c_term_comparison.items()):
                    values = [data.get(residue, 0) for residue in sorted(all_c_residues)]
                    ax_c_term.bar(x_pos + idx * width, values, width, 
                                 label=condition, alpha=0.8)
                
                ax_c_term.set_xlabel('C-terminal Cleavage Residue')
                ax_c_term.set_ylabel('Total Intensity')
                ax_c_term.set_title('C-terminal Cleavage Comparison')
                ax_c_term.set_xticks(x_pos + width)
                ax_c_term.set_xticklabels(sorted(all_c_residues))
                ax_c_term.legend(fontsize=8)
        
        # Plot total intensities
        conditions_list = list(total_intensities.keys())
        totals_list = list(total_intensities.values())
        
        bars = ax_total.bar(conditions_list, totals_list, color=['skyblue', 'lightgreen', 'salmon'])
        ax_total.set_xlabel('Condition')
        ax_total.set_ylabel('Total Intensity')
        ax_total.set_title('Total Intensity Comparison')
        ax_total.tick_params(axis='x', rotation=45)
        
        # Add value labels on bars
        for bar, value in zip(bars, totals_list):
            ax_total.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                         f'{value/1e6:.1f}M', ha='center', va='bottom', fontsize=8)
        
        # Bottom row: Summary statistics table
        ax_table = fig.add_subplot(gs[2, :])
        ax_table.axis('off')
        
        # Create summary table data
        table_data = []
        headers = ['Condition', 'Sequences', 'Max Intensity', 'Mean Intensity', 'Positions w/ Data']
        
        for condition_name, raw_data in all_data.items():
            sequences = raw_data['sequences']
            all_intensities = [i for seq in sequences for i in seq['intensities'] if i > 0]
            
            # Count positions with data
            reference = raw_data['reference']
            position_has_data = [False] * len(reference)
            for seq_data in sequences:
                clean_seq = seq_data['clean']
                if clean_seq in reference and any(i > 0 for i in seq_data['intensities']):
                    start_pos = reference.index(clean_seq)
                    end_pos = start_pos + len(clean_seq)
                    for pos in range(start_pos, end_pos):
                        position_has_data[pos] = True
            
            table_data.append([
                condition_name,
                str(len(sequences)),
                f"{max(all_intensities)/1e6:.1f}M" if all_intensities else "0",
                f"{np.mean(all_intensities)/1e6:.1f}M" if all_intensities else "0",
                f"{sum(position_has_data)}/{len(reference)}"
            ])
        
        # Create table
        table = ax_table.table(cellText=table_data, colLabels=headers,
                              cellLoc='center', loc='center')
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1, 2)
        
        # Add colorbar for the heatmaps
        cbar_ax = fig.add_axes([0.92, 0.55, 0.02, 0.3])
        cbar = plt.colorbar(im, cax=cbar_ax)
        cbar.set_label('Log10(Intensity + 1)', rotation=270, labelpad=15)
        
        plt.suptitle('Comprehensive Cleavage Analysis Report', fontsize=16, y=0.95)
        plt.tight_layout()
        plt.subplots_adjust(top=0.92, right=0.9)
        
        # Save the report
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"✓ Comprehensive report saved to: {output_path}")
        
        # Print summary
        print(f"✓ Report includes:")
        print(f"  - {len(all_data)} conditions compared")
        print(f"  - Positional intensity heatmaps")
        print(f"  - Cleavage pattern analysis")
        print(f"  - Summary statistics")
        
        return fig


# ============================================================================
# EXECUTION EXAMPLES
# ============================================================================

def example_single_sheet():
    """Process a single worksheet"""
    mapper = AdvancedCleavageMapper('example_data_converted.xlsx')
    
    sample_names = [
        'AspN_Fxn2', 'AspN_Fxn3', 'AspN_Fxn4', 
        'AspN_Fxn5', 'AspN_Fxn6', 'AspN_Fxn7', 'AspN_Fxn8'
    ]
    
    mapper.process('500 mgd glucose', '500 mgd PROCESSED', sample_names)
    mapper.save()

def example_batch_processing():
    """Process multiple worksheets"""
    mapper = AdvancedCleavageMapper('hu_stem_cell_3_glucose_conditions_edit3.xlsx')
    
    sample_names = [
        'AspN_Fxn2', 'AspN_Fxn3', 'AspN_Fxn4',
        'AspN_Fxn5', 'AspN_Fxn6', 'AspN_Fxn7', 'AspN_Fxn8'
    ]
    
    worksheets = [
        ('100 mgd glucose', '100 mgd PROCESSED'),
        ('200 mgd glucose', '200 mgd PROCESSED'),
        ('500 mgd glucose', '500 mgd PROCESSED')
    ]
    
    for input_ws, output_ws in worksheets:
        mapper.process(input_ws, output_ws, sample_names)
    
    mapper.save('processed_output.xlsx')

def example_with_validation():
    """Process with validation checks"""
    mapper = AdvancedCleavageMapper('hu_stem_cell_3_glucose_conditions_edit3.xlsx')
    
    try:
        raw_data = mapper.parse_raw_worksheet('500 mgd glucose')
        
        # Validate
        assert len(raw_data['sequences']) > 0, "No sequences found"
        assert raw_data['reference'], "No reference sequence"
        
        # Check cleavage notation
        for seq in raw_data['sequences']:
            assert seq['left_cleavage'], f"Missing left cleavage: {seq['original']}"
            assert seq['right_cleavage'], f"Missing right cleavage: {seq['original']}"
        
        # Process
        mapper.process('500 mgd glucose', '500 mgd VALIDATED')
        mapper.save()
        
        print("\nValidation passed")
        
    except AssertionError as e:
        print(f"\nValidation failed: {e}")


if __name__ == "__main__":
    # Run desired example
    example_single_sheet()